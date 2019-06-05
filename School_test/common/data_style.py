# -*- coding:utf-8 -*-
# Author:wangjian

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Font
import csv,os,time

#wj_data路径
now = time.strftime("%Y-%m-%d_%H-%M-%S",time.localtime(time.time()))
csv_file = now+"data.csv"
path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
data_path = os.path.join(os.path.join(path, "wj_data"),csv_file)

class Excle(object):
    """Excel文件的操作方法"""
    def __init__(self,test_path):
        self.font = Font(color=None)
        self.colorDict = {'red':'FFFF3030','green':'FF008B00'}
        self.test_path = test_path
        self.wb = load_workbook(self.test_path)
        self.ws = self.wb.active

    def rename(self,new_name):
        #表格重命名
        self.ws.title = new_name
        self.wb.save(self.test_path)

    def GetSheetName(self):
        #获取所有表格名称
        return self.wb.get_sheet_names()

    def GetSheetByName(self,sheet_name):
        #通过表格名称获取表格
        self.ws = self.wb.get_sheet_by_name(sheet_name)
        return self.ws

    def GetCurrentSheetName(self):
        #获取当前表格名称
        return self.ws.title

    def GetCellContent(self,row_num,col_num):
        #获取单元格内容
        return self.ws.cell(row=row_num,column=col_num).value

    def WriteCellContent(self,row_num,col_num,content):
        #指定的单元格里面写入内容
        self.ws.cell(row=row_num,column=col_num).value=content
        self.wb.save(self.test_path)

    def GetMaxRow(self):
        #获取最大行号
        return self.ws.max_row

    def GetMaxColumn(self):
        #获取最大列号
        return self.ws.max_column

class Csv(object):
    """CSV文件的操作方法"""
    def __init__(self,filename):
        self.filename = filename

    def get_row(self,name):
        #获取列的数据
        try:
            with open(self.filename,"r",newline="",encoding="utf-8") as csvfile:
                values = csv.DictReader(csvfile)
                rows = [row[name] for row in values]
                return rows
        except Exception as e:
            raise e
    def get_col(self,name):
        #获取行的数据
        try:
            with open(self.filename,"r",newline="",encoding="utf-8") as csvfile:
                values = csv.reader(csvfile)
                cols = [col[name] for col in values]
                return cols
        except Exception as e:
            raise e
    def write_row(self,fieldnames,*datas):
        #写入列的数据
        try:
            with open(self.filename,"w",newline="",encoding="utf-8") as csvfile:
                writer = csv.DictWriter(csvfile,fieldnames=fieldnames)
                writer.writeheader()
                dic = {}
                for data in datas:
                    for index,i in enumerate(fieldnames):
                        dic[i] = data[index]
                    writer.writerow(dic)
        except Exception as e:
            raise e
    def write_col(self,datas):
        #写入行的数据
        try:
            with open(self.filename,'w',newline="",encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile,dialect='excel')
                for data in datas:
                    writer.writerow(data)
        except Exception as e:
            raise e
if __name__ == "__main__":
    c =Csv(data_path)
    # c.write_row(["first_name","last_name"],(["wang","jian"],["li","kang"],["dsad","dsad"]))
    # c.write_col([["a","b","c","d","e"],[1,2,3,4,5]])

